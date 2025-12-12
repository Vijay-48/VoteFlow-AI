"""
VoteFlow AI - Backend Server
============================
FastAPI backend for campaign execution with WebSocket support.
Includes security measures: input validation, rate limiting, sanitization.
Multi-user support with payment tracking and quota management.
"""

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, HTTPException, BackgroundTasks, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel, validator
import asyncio
import json
import os
import re
import html
from datetime import datetime
from typing import Dict, List, Optional
import uuid
from collections import defaultdict
import time
import razorpay
import hmac
import hashlib

# Import our modules
from campaign_runner import CampaignRunner
from pdf_extractor import PDFExtractor
from database import (
    init_db, get_db, get_user_by_clerk_id, get_user_quota, 
    decrement_quota, add_quota, get_user_chrome_profile,
    PACKAGE_LIMITS, PackageType, User, Payment, PaymentStatus, SessionLocal
)

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# Razorpay client
RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID", "")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET", "")
razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET)) if RAZORPAY_KEY_ID else None

app = FastAPI(
    title="VoteFlow AI Backend",
    description="Campaign execution and PDF extraction API with multi-user support",
    version="2.0.0"
)

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    init_db()
    print("[DB] ✅ Database initialized")

# CORS for React frontend
app.add_middleware(
    CORSMiddleware,
    # allow_origins=origins,  <-- Replaced with regex below
    allow_origin_regex=".*",  # Allow ALL origins (Nuclear option for debugging)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =============================================================================
# Security: Rate Limiting
# =============================================================================

class RateLimiter:
    def __init__(self, max_requests: int = 10, time_window: int = 60):
        self.max_requests = max_requests
        self.time_window = time_window
        self.requests = defaultdict(list)
    
    def is_allowed(self, client_ip: str) -> bool:
        now = time.time()
        # Clean old requests
        self.requests[client_ip] = [
            req_time for req_time in self.requests[client_ip]
            if now - req_time < self.time_window
        ]
        # Check rate limit
        if len(self.requests[client_ip]) >= self.max_requests:
            return False
        self.requests[client_ip].append(now)
        return True

rate_limiter = RateLimiter(max_requests=30, time_window=60)  # 30 requests per minute

# =============================================================================
# Security: Input Sanitization
# =============================================================================

def sanitize_string(input_str: str, max_length: int = 1000) -> str:
    """Sanitize string input to prevent XSS and injection attacks."""
    if not input_str:
        return ""
    
    # Truncate to max length
    input_str = input_str[:max_length]
    
    # HTML escape
    input_str = html.escape(input_str)
    
    # Remove potential SQL injection patterns
    dangerous_patterns = [
        r"(\-\-)", r"(\/\*)", r"(\*\/)", r"(;)",
        r"(DROP\s+TABLE)", r"(DELETE\s+FROM)", r"(INSERT\s+INTO)",
        r"(UPDATE\s+\w+\s+SET)", r"(UNION\s+SELECT)"
    ]
    for pattern in dangerous_patterns:
        input_str = re.sub(pattern, "", input_str, flags=re.IGNORECASE)
    
    return input_str

def validate_phone_number(phone: str) -> bool:
    """Validate phone number format."""
    clean_phone = re.sub(r'\D', '', phone)
    return len(clean_phone) == 10 or len(clean_phone) == 12

# =============================================================================
# Pydantic Models for Request Validation
# =============================================================================

class CampaignStartRequest(BaseModel):
    user_id: Optional[str] = None  # Clerk user ID for quota and per-user profile
    message_template: Optional[str] = None
    voters_data: Optional[List[dict]] = None
    excel_file: Optional[str] = None
    plan_id: Optional[str] = None
    max_messages: Optional[int] = 500
    
    @validator('message_template')
    def sanitize_message(cls, v):
        if v:
            return sanitize_string(v, max_length=2000)
        return v
    
    @validator('max_messages')
    def validate_max_messages(cls, v):
        # Enforce plan limits
        if v > 10000:
            return 10000
        if v < 1:
            return 1
        return v

# =============================================================================
# Global State
# =============================================================================

active_campaigns: Dict[str, CampaignRunner] = {}
websocket_connections: Dict[str, List[WebSocket]] = {}

# =============================================================================
# WebSocket Manager
# =============================================================================

class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, websocket: WebSocket, campaign_id: str):
        await websocket.accept()
        if campaign_id not in self.active_connections:
            self.active_connections[campaign_id] = []
        self.active_connections[campaign_id].append(websocket)

    def disconnect(self, websocket: WebSocket, campaign_id: str):
        if campaign_id in self.active_connections:
            if websocket in self.active_connections[campaign_id]:
                self.active_connections[campaign_id].remove(websocket)

    async def broadcast(self, campaign_id: str, message: dict):
        if campaign_id in self.active_connections:
            for connection in self.active_connections[campaign_id]:
                try:
                    await connection.send_json(message)
                except:
                    pass

manager = ConnectionManager()

# =============================================================================
# REST API Endpoints
# =============================================================================

@app.get("/")
async def root():
    return {
        "status": "online",
        "service": "VoteFlow AI Backend",
        "version": "1.0.0",
        "security": "enabled",
        "endpoints": {
            "health": "/health",
            "extract_pdf": "POST /api/extract-pdf",
            "start_campaign": "POST /api/campaign/start",
            "stop_campaign": "POST /api/campaign/{id}/stop",
            "campaign_status": "GET /api/campaign/{id}/status",
            "websocket": "WS /ws/campaign/{id}"
        }
    }

@app.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

# =============================================================================
# PDF Extraction Endpoint (with security)
# =============================================================================

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

@app.post("/api/extract-pdf")
async def extract_pdf(request: Request, file: UploadFile = File(...)):
    """
    Upload a PDF with voter information (including handwritten mobile numbers).
    Uses Gemini Vision AI to extract all data including OCR for handwritten text.
    
    Security: Rate limited, file size limited, file type validated.
    """
    # Rate limiting
    client_ip = request.client.host
    if not rate_limiter.is_allowed(client_ip):
        raise HTTPException(status_code=429, detail="Too many requests. Please wait.")
    
    # File type validation
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")
    
    # Read file content
    content = await file.read()
    
    # File size validation
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size must be less than 10MB")
    
    try:
        # Save uploaded file temporarily
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        
        # Secure filename
        safe_filename = re.sub(r'[^\w\-.]', '_', file.filename)
        file_path = os.path.join(upload_dir, f"{uuid.uuid4()}_{safe_filename}")
        
        with open(file_path, "wb") as f:
            f.write(content)
        
        # Extract data using Gemini Vision
        extractor = PDFExtractor()
        result = await extractor.extract_voters(file_path)
        
        # Clean up temp file
        os.remove(file_path)
        
        # Sanitize extracted data
        for voter in result.get('voters', []):
            if 'NAME' in voter:
                voter['NAME'] = sanitize_string(voter['NAME'], 100)
            if 'MOBILE' in voter:
                voter['MOBILE'] = re.sub(r'\D', '', str(voter['MOBILE']))[:12]
        
        return JSONResponse({
            "success": True,
            "message": f"Extracted {len(result['voters'])} voters",
            "data": result
        })
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# Excel Upload Endpoint (Quick Upload - No AI needed)
# =============================================================================

@app.post("/api/upload-excel")
async def upload_excel(request: Request, file: UploadFile = File(...)):
    """
    Upload an Excel file with pre-filled voter data (NAME and PHONE columns).
    This is a quick upload option that doesn't require AI processing.
    """
    # Rate limiting
    client_ip = request.client.host
    if not rate_limiter.is_allowed(client_ip):
        raise HTTPException(status_code=429, detail="Too many requests. Please wait.")
    
    # File type validation
    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are accepted")
    
    # Read file content
    content = await file.read()
    
    # File size validation
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size must be less than 10MB")
    
    try:
        import openpyxl
        from io import BytesIO
        
        # Load workbook
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        
        voters = []
        headers = []
        name_col = None
        phone_col = None
        
        # Find header row and column mappings
        for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_lower = str(cell).lower().strip()
                    if any(n in cell_lower for n in ['name', 'voter', 'పేరు']):
                        name_col = col_idx
                        headers = row
                    if any(p in cell_lower for p in ['phone', 'mobile', 'number', 'ఫోన్', 'mob']):
                        phone_col = col_idx
            if name_col is not None:
                break
        
        if name_col is None:
            # Default to first two columns
            name_col = 0
            phone_col = 1
        
        if phone_col is None:
            phone_col = 1
        
        # Extract data
        start_row = 2 if headers else 1
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row[name_col] and len(row) > phone_col:
                name = str(row[name_col]).strip()
                phone_raw = str(row[phone_col] if row[phone_col] else '').strip()
                phone = re.sub(r'\D', '', phone_raw)
                
                # Validate phone
                if len(phone) == 10 and phone[0] in '6789':
                    voters.append({
                        'NAME': sanitize_string(name, 100),
                        'MOBILE': phone
                    })
                elif len(phone) == 12 and phone.startswith('91'):
                    voters.append({
                        'NAME': sanitize_string(name, 100),
                        'MOBILE': phone[2:]  # Remove +91
                    })
        
        return JSONResponse({
            "success": True,
            "message": f"Loaded {len(voters)} voters from Excel",
            "voters": voters
        })
        
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# Payment Verification Endpoint
# =============================================================================

# Store verified payment IDs to prevent reuse
verified_payments = set()

class PaymentVerifyRequest(BaseModel):
    payment_id: str
    plan_id: str
    amount: int

@app.post("/api/verify-payment")
async def verify_payment(request: Request, body: PaymentVerifyRequest):
    """
    Verify a Razorpay payment ID.
    
    For production, integrate with Razorpay API to verify:
    https://razorpay.com/docs/api/payments/#fetch-a-payment
    
    Current implementation: Validates format and prevents reuse.
    """
    # Rate limiting
    client_ip = request.client.host
    if not rate_limiter.is_allowed(client_ip):
        raise HTTPException(status_code=429, detail="Too many requests. Please wait.")
    
    payment_id = body.payment_id.strip()
    
    # Validate payment ID format (Razorpay format: pay_xxxxxxxxxxxxxx)
    if not re.match(r'^pay_[a-zA-Z0-9]{10,30}$', payment_id):
        return {"verified": False, "message": "Invalid Payment ID format"}
    
    # Check if already used
    if payment_id in verified_payments:
        return {"verified": False, "message": "This Payment ID has already been used"}
    
    # For production: Add Razorpay API verification here
    # import razorpay
    # client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
    # payment = client.payment.fetch(payment_id)
    # if payment['status'] != 'captured' or payment['amount'] != body.amount * 100:
    #     return {"verified": False, "message": "Payment not found or amount mismatch"}
    
    # Mark as used
    verified_payments.add(payment_id)
    
    return {
        "verified": True,
        "message": "Payment verified successfully",
        "payment_id": payment_id
    }

# =============================================================================
# Campaign Control Endpoints (with validation)
# =============================================================================

@app.post("/api/campaign/start")
async def start_campaign(
    request: Request,
    body: CampaignStartRequest,
    background_tasks: BackgroundTasks
):
    """
    Start a new campaign execution.
    
    Security: Rate limited, input sanitized, message length limited.
    """
    # Rate limiting
    client_ip = request.client.host
    if not rate_limiter.is_allowed(client_ip):
        raise HTTPException(status_code=429, detail="Too many requests. Please wait.")
    
    campaign_id = str(uuid.uuid4())[:8]
    
    try:
        # Validate and sanitize voters data
        voters_data = body.voters_data or []
        sanitized_voters = []
        
        for voter in voters_data[:body.max_messages]:  # Limit by plan
            if isinstance(voter, dict):
                sanitized_voter = {
                    'NAME': sanitize_string(voter.get('NAME', ''), 100),
                    'MOBILE': re.sub(r'\D', '', str(voter.get('MOBILE', '')))[:12]
                }
                # Only include if valid mobile
                if validate_phone_number(sanitized_voter['MOBILE']):
                    sanitized_voters.append(sanitized_voter)
        
        # Check user quota if user_id is provided
        user_quota = 0
        if body.user_id:
            db = SessionLocal()
            try:
                quota_info = get_user_quota(db, body.user_id)
                user_quota = quota_info.get('messages_remaining', 0)
                if user_quota <= 0:
                    raise HTTPException(status_code=403, detail="No messages remaining. Please purchase a package.")
                # Limit voters to quota
                if len(sanitized_voters) > user_quota:
                    sanitized_voters = sanitized_voters[:user_quota]
            finally:
                db.close()
        
        runner = CampaignRunner(
            campaign_id=campaign_id,
            user_id=body.user_id,
            message_template=body.message_template,
            broadcast_callback=lambda msg: asyncio.create_task(
                manager.broadcast(campaign_id, msg)
            ),
            quota_remaining=user_quota
        )
        
        active_campaigns[campaign_id] = runner
        
        # Start campaign in background
        background_tasks.add_task(
            runner.execute_campaign,
            excel_file=body.excel_file,
            voters_data=sanitized_voters if sanitized_voters else body.voters_data
        )
        
        return {
            "success": True,
            "campaign_id": campaign_id,
            "voters_count": len(sanitized_voters) if sanitized_voters else 0,
            "message": "Campaign started. Connect to WebSocket for live updates.",
            "websocket_url": f"ws://localhost:8000/ws/campaign/{campaign_id}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/campaign/{campaign_id}/stop")
async def stop_campaign(campaign_id: str):
    """Stop a running campaign."""
    # Validate campaign_id format
    if not re.match(r'^[a-f0-9\-]{8}$', campaign_id):
        raise HTTPException(status_code=400, detail="Invalid campaign ID format")
    
    if campaign_id not in active_campaigns:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    runner = active_campaigns[campaign_id]
    runner.stop()
    
    return {"success": True, "message": f"Campaign {campaign_id} stopped"}

@app.get("/api/campaign/{campaign_id}/status")
async def campaign_status(campaign_id: str):
    """Get current status of a campaign."""
    # Validate campaign_id format
    if not re.match(r'^[a-f0-9\-]{8}$', campaign_id):
        raise HTTPException(status_code=400, detail="Invalid campaign ID format")
    
    if campaign_id not in active_campaigns:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    runner = active_campaigns[campaign_id]
    return runner.get_status()

# =============================================================================
# WebSocket for Real-time Logs
# =============================================================================

@app.websocket("/ws/campaign/{campaign_id}")
async def websocket_endpoint(websocket: WebSocket, campaign_id: str):
    """
    WebSocket endpoint for real-time campaign logs.
    Connect to receive live updates during campaign execution.
    """
    # Validate campaign_id format
    if not re.match(r'^[a-f0-9\-]{8}$', campaign_id):
        await websocket.close(code=4000)
        return
    
    await manager.connect(websocket, campaign_id)
    
    try:
        # Send initial connection message
        await websocket.send_json({
            "type": "connected",
            "campaign_id": campaign_id,
            "message": "Connected to campaign log stream"
        })
        
        # Keep connection alive and handle any incoming messages
        while True:
            try:
                data = await asyncio.wait_for(websocket.receive_text(), timeout=30)
                # Handle ping/pong or commands
                if data == "ping":
                    await websocket.send_text("pong")
            except asyncio.TimeoutError:
                # Send heartbeat
                await websocket.send_json({"type": "heartbeat"})
                
    except WebSocketDisconnect:
        manager.disconnect(websocket, campaign_id)

# =============================================================================
# WhatsApp Sandbox Status
# =============================================================================

@app.get("/api/whatsapp/status")
async def whatsapp_status():
    """Check if WhatsApp Web session is active."""
    return {
        "connected": False,
        "message": "Start a campaign to initialize WhatsApp Web connection",
        "instructions": [
            "1. Click 'Start Campaign' to open WhatsApp Web",
            "2. Scan the QR code with your phone",
            "3. Once connected, messages will be sent automatically"
        ]
    }

@app.get("/api/whatsapp/qr")
async def get_whatsapp_qr():
    """Get the WhatsApp QR code screenshot for remote scanning."""
    from fastapi.responses import FileResponse
    
    qr_path = os.path.join(os.path.dirname(__file__), "uploads", "qr_code.png")
    
    if os.path.exists(qr_path):
        return FileResponse(
            qr_path, 
            media_type="image/png",
            headers={"Cache-Control": "no-cache, no-store, must-revalidate"}
        )
    else:
        raise HTTPException(
            status_code=404, 
            detail="QR code not available. Start a campaign first to generate the QR code."
        )

@app.get("/api/whatsapp/vnc-url/{user_id}")
async def get_vnc_url(user_id: str, request: Request):
    """Get the noVNC URL for a user to link their WhatsApp account."""
    # Get the host from the request
    host = request.headers.get("host", "localhost")
    base_url = host.split(":")[0]  # Remove port if present
    
    # VNC port (mapped in docker-compose)
    vnc_port = 6080
    
    # For production, use the actual domain
    if "api.techmans.me" in host:
        vnc_url = f"https://api.techmans.me:{vnc_port}/vnc.html?autoconnect=true"
    else:
        vnc_url = f"http://{base_url}:{vnc_port}/vnc.html?autoconnect=true"
    
    return {
        "success": True,
        "vnc_url": vnc_url,
        "user_id": user_id,
        "instructions": [
            "1. Click the VNC URL to open Chrome in your browser",
            "2. Navigate to web.whatsapp.com in the Chrome window",
            "3. Scan the QR code with your phone",
            "4. Once connected, close the VNC window"
        ]
    }

@app.get("/api/whatsapp/status/{user_id}")
async def get_whatsapp_status(user_id: str):
    """Check if a user has linked their WhatsApp account."""
    db = SessionLocal()
    try:
        user = get_user_by_clerk_id(db, user_id)
        
        # Check if profile directory exists and has session data
        profile_path = get_user_chrome_profile(user_id)
        session_file = os.path.join(profile_path, "Default", "IndexedDB")
        has_session = os.path.exists(session_file)
        
        return {
            "success": True,
            "user_id": user_id,
            "whatsapp_linked": user.whatsapp_linked or has_session,
            "whatsapp_phone": user.whatsapp_phone
        }
    finally:
        db.close()

@app.post("/api/whatsapp/mark-linked/{user_id}")
async def mark_whatsapp_linked(user_id: str, phone: Optional[str] = None):
    """Mark a user's WhatsApp as linked (called after successful QR scan)."""
    db = SessionLocal()
    try:
        user = get_user_by_clerk_id(db, user_id)
        user.whatsapp_linked = True
        if phone:
            user.whatsapp_phone = phone
        user.chrome_profile_path = get_user_chrome_profile(user_id)
        db.commit()
        
        return {
            "success": True,
            "message": "WhatsApp linked successfully",
            "user_id": user_id
        }
    finally:
        db.close()

# =============================================================================
# Payment API Endpoints
# =============================================================================

class CreateOrderRequest(BaseModel):
    """Request model for creating a payment order."""
    user_id: str  # Clerk user ID
    package_type: str  # starter, growth, enterprise
    email: Optional[str] = None
    name: Optional[str] = None

class VerifyPaymentRequest(BaseModel):
    """Request model for verifying payment."""
    user_id: str
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str

@app.post("/api/payment/create-order")
async def create_payment_order(request: CreateOrderRequest):
    """Create a Razorpay order for package purchase."""
    if not razorpay_client:
        raise HTTPException(status_code=500, detail="Payment gateway not configured")
    
    # Validate package type
    try:
        pkg = PackageType(request.package_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid package type: {request.package_type}")
    
    package_info = PACKAGE_LIMITS[pkg]
    amount_paise = int(package_info["price"] * 100)  # Razorpay uses paise
    
    try:
        # Create Razorpay order
        order_data = {
            "amount": amount_paise,
            "currency": "INR",
            "receipt": f"order_{uuid.uuid4().hex[:8]}",
            "notes": {
                "user_id": request.user_id,
                "package_type": request.package_type,
                "messages": package_info["messages"]
            }
        }
        order = razorpay_client.order.create(data=order_data)
        
        # Save order to database
        db = SessionLocal()
        try:
            user = get_user_by_clerk_id(db, request.user_id)
            if request.email:
                user.email = request.email
            if request.name:
                user.name = request.name
            
            payment = Payment(
                user_id=user.id,
                amount=package_info["price"],
                package_type=request.package_type,
                messages_purchased=package_info["messages"],
                razorpay_order_id=order["id"],
                status=PaymentStatus.PENDING.value
            )
            db.add(payment)
            db.commit()
        finally:
            db.close()
        
        return {
            "success": True,
            "order_id": order["id"],
            "amount": package_info["price"],
            "currency": "INR",
            "key_id": RAZORPAY_KEY_ID,
            "package": {
                "type": request.package_type,
                "messages": package_info["messages"]
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create order: {str(e)}")

@app.post("/api/payment/verify")
async def verify_payment(request: VerifyPaymentRequest):
    """Verify Razorpay payment and add quota to user."""
    if not razorpay_client:
        raise HTTPException(status_code=500, detail="Payment gateway not configured")
    
    # Verify signature
    try:
        params_dict = {
            "razorpay_order_id": request.razorpay_order_id,
            "razorpay_payment_id": request.razorpay_payment_id,
            "razorpay_signature": request.razorpay_signature
        }
        razorpay_client.utility.verify_payment_signature(params_dict)
    except Exception:
        raise HTTPException(status_code=400, detail="Payment verification failed")
    
    # Update database
    db = SessionLocal()
    try:
        payment = db.query(Payment).filter(Payment.razorpay_order_id == request.razorpay_order_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="Order not found")
        
        payment.razorpay_payment_id = request.razorpay_payment_id
        payment.razorpay_signature = request.razorpay_signature
        payment.status = PaymentStatus.COMPLETED.value
        payment.completed_at = datetime.utcnow()
        
        # Add quota to user
        user = db.query(User).filter(User.id == payment.user_id).first()
        user.messages_remaining += payment.messages_purchased
        user.package_type = payment.package_type
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Payment verified! Added {payment.messages_purchased} messages.",
            "quota": {
                "messages_remaining": user.messages_remaining,
                "package_type": user.package_type
            }
        }
    finally:
        db.close()

@app.get("/api/user/quota/{user_id}")
async def get_user_quota_endpoint(user_id: str):
    """Get user's remaining message quota."""
    db = SessionLocal()
    try:
        quota = get_user_quota(db, user_id)
        return {
            "success": True,
            **quota
        }
    finally:
        db.close()

@app.get("/api/user/profile/{user_id}")
async def get_user_profile(user_id: str):
    """Get user profile including WhatsApp status."""
    db = SessionLocal()
    try:
        user = get_user_by_clerk_id(db, user_id)
        return {
            "success": True,
            "user": {
                "id": user.clerk_user_id,
                "email": user.email,
                "name": user.name,
                "package_type": user.package_type,
                "messages_remaining": user.messages_remaining,
                "total_sent": user.total_messages_sent,
                "whatsapp_linked": user.whatsapp_linked
            }
        }
    finally:
        db.close()

# =============================================================================
# Entry Point
# =============================================================================

if __name__ == "__main__":
    import uvicorn
    print("🚀 Starting VoteFlow AI Backend...")
    print("🔐 Security: Rate limiting & input validation enabled")
    print("💳 Payments: Razorpay integration active")
    print("📡 API: http://localhost:8000")
    print("📖 Docs: http://localhost:8000/docs")
    uvicorn.run(app, host="0.0.0.0", port=8000)
